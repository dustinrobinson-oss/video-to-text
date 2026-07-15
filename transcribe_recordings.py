"""
transcribe_recordings.py
========================
RealWork Labs, Shared Transcription Automation

Three input modes:
  1. Local files:  drop audio/video into the input folder
  2. URLs:         pass --url flags or put URLs in links.txt
  3. Excel:        pass --xlsx with a spreadsheet of URLs (e.g. Zoom recording links)

Downloads audio via yt-dlp, transcribes with OpenAI Whisper, saves .txt
transcripts, and archives originals. Smart dedup prevents reprocessing.

Setup (one time):
    pip install openai-whisper yt-dlp openpyxl
    # Also needs ffmpeg:
    #   Windows: choco install ffmpeg  (Admin PowerShell)
    #   macOS:   brew install ffmpeg
    #   Linux:   sudo apt install ffmpeg

Usage:
    # Local files
    python transcribe_recordings.py --input /path/to/recordings --output /path/to/transcripts

    # Single URL
    python transcribe_recordings.py --url "https://youtube.com/watch?v=..." --output /path/to/transcripts

    # Excel file with Zoom recording links (uses browser cookies for auth)
    python transcribe_recordings.py --xlsx calls.xlsx --output /path/to/transcripts

    # Excel with custom column names
    python transcribe_recordings.py --xlsx calls.xlsx --url-column "All Recording Links" --name-columns "Account" "Touchplan"

Defaults:
    Input:   ~/Documents/Claude/sources/recordings/
    Output:  ~/Documents/Claude/sources/transcripts/
    Archive: ~/Documents/Claude/sources/recordings/Archive/
    Model:   medium
"""

import argparse
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

AUDIO_VIDEO_EXTENSIONS = {
    ".mp4", ".mov", ".avi", ".mkv", ".webm",
    ".mp3", ".wav", ".m4a", ".flac", ".ogg", ".aac",
}

DEFAULT_INPUT = Path.home() / "Documents" / "Claude" / "sources" / "recordings"
DEFAULT_OUTPUT = Path.home() / "Documents" / "Claude" / "sources" / "transcripts"

LINKS_FILENAME = "links.txt"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Transcribe audio/video files, URLs, and Excel link lists using Whisper."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT,
                        help="Folder containing audio/video files to transcribe")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help="Folder to save .txt transcripts")
    parser.add_argument("--model", default="medium",
                        choices=["tiny", "tiny.en", "base", "base.en", "small",
                                 "small.en", "medium", "medium.en", "large"],
                        help="Whisper model size (default: medium)")
    parser.add_argument("--language", default="en",
                        help="Language code (default: en)")
    parser.add_argument("--keep-originals", action="store_true",
                        help="Don't move originals to Archive after transcribing")
    parser.add_argument("--url", action="append", default=[],
                        help="URL to transcribe (repeatable)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Excel file containing URLs to transcribe")
    parser.add_argument("--url-column", default="Recording Link",
                        help="Column name in Excel containing URLs (default: Recording Link)")
    parser.add_argument("--name-columns", nargs="+",
                        default=["Account", "Touchplan"],
                        help="Column names to build transcript filename (default: Account Touchplan)")
    parser.add_argument("--cookies-from-browser", default=None,
                        help="Browser to pull cookies from for authenticated downloads (chrome, firefox, edge)")
    return parser.parse_args()


# ---------------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------------

def find_media_files(input_dir: Path) -> list[Path]:
    """Find all audio/video files in input_dir (not subdirectories)."""
    if not input_dir.is_dir():
        return []
    return sorted([
        f for f in input_dir.iterdir()
        if f.is_file() and f.suffix.lower() in AUDIO_VIDEO_EXTENSIONS
    ])


def get_existing_transcripts(output_dir: Path) -> set[str]:
    """Get stems of all existing .txt transcripts (lowercase for matching)."""
    if not output_dir.exists():
        return set()
    return {f.stem.lower() for f in output_dir.iterdir() if f.suffix.lower() == ".txt"}


def transcript_exists(filename_stem: str, existing: set[str]) -> bool:
    """Check if a transcript already exists for this recording."""
    stem_lower = filename_stem.lower()
    if stem_lower in existing:
        return True
    for t in existing:
        if t.startswith(stem_lower):
            return True
    return False


def safe_output_path(output_dir: Path, stem: str, suffix: str = ".txt") -> Path:
    """Return a unique output path, appending date if file already exists."""
    candidate = output_dir / f"{stem}{suffix}"
    if not candidate.exists():
        return candidate
    dated = output_dir / f"{stem}_{datetime.now().strftime('%Y-%m-%d')}{suffix}"
    if not dated.exists():
        return dated
    return output_dir / f"{stem}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}{suffix}"


def safe_archive_path(archive_dir: Path, filename: str) -> Path:
    """Return a unique archive path."""
    candidate = archive_dir / filename
    if not candidate.exists():
        return candidate
    stem = Path(filename).stem
    ext = Path(filename).suffix
    return archive_dir / f"{stem}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}{ext}"


def sanitize_filename(title: str) -> str:
    """Convert a string into a safe filename stem."""
    clean = re.sub(r'[<>:"/\\|?*]', '', title)
    clean = re.sub(r'\s+', '_', clean.strip())
    clean = clean.strip('_.')
    if len(clean) > 120:
        clean = clean[:120].rstrip('_')
    return clean if clean else "untitled"


# ---------------------------------------------------------------------------
# links.txt support
# ---------------------------------------------------------------------------

def read_links_file(input_dir: Path) -> list[str]:
    """Read URLs from links.txt in the input directory, one per line."""
    links_path = input_dir / LINKS_FILENAME
    if not links_path.exists():
        return []
    urls = []
    with open(links_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                urls.append(line)
    return urls


def archive_links_file(input_dir: Path, archive_dir: Path) -> None:
    """Move processed links.txt to Archive with a timestamp."""
    links_path = input_dir / LINKS_FILENAME
    if links_path.exists():
        dest = safe_archive_path(
            archive_dir,
            f"links_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.txt"
        )
        shutil.move(str(links_path), str(dest))
        print(f"  Archived links.txt to: {dest.name}")


# ---------------------------------------------------------------------------
# Excel support
# ---------------------------------------------------------------------------

def read_xlsx_urls(xlsx_path: Path, url_column: str, name_columns: list[str]) -> list[dict]:
    """Read URLs and metadata from an Excel file.

    Returns a list of dicts with keys: url, name, metadata (dict of all row values).
    Handles cells with multiple newline-separated URLs by splitting them into
    separate entries.
    """
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Map header names to column indices
    headers = [cell.value for cell in ws[1]]
    header_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            header_map[str(h).strip()] = i

    # Validate required columns exist
    if url_column not in header_map:
        print(f"Error: Column '{url_column}' not found in Excel file.")
        print(f"Available columns: {[h for h in header_map.keys()]}")
        sys.exit(1)

    url_idx = header_map[url_column]
    name_indices = [header_map[c] for c in name_columns if c in header_map]

    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_url = row[url_idx]
        if not raw_url or not str(raw_url).strip():
            continue

        # Build metadata dict for the row
        metadata = {}
        for col_name, col_idx in header_map.items():
            val = row[col_idx]
            if val is not None:
                metadata[col_name] = str(val).strip()

        # Build filename from name columns
        name_parts = []
        for idx in name_indices:
            val = row[idx]
            if val is not None:
                name_parts.append(str(val).strip())
        base_name = sanitize_filename("_".join(name_parts)) if name_parts else None

        # Split cell on newlines (handles multi-URL cells)
        url_strings = str(raw_url).strip().split("\n")
        for url_num, url in enumerate(url_strings):
            url = url.strip()
            if not url or not url.startswith("http"):
                continue
            # Append call number if multiple URLs from same row
            if len(url_strings) > 1 and base_name:
                entry_name = f"{base_name}_call{url_num + 1}"
            else:
                entry_name = base_name
            entries.append({
                "url": url,
                "name": entry_name,
                "metadata": metadata,
            })

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# URL download
# ---------------------------------------------------------------------------

def download_url_audio(url: str, download_dir: Path, filename: str = None,
                       cookies_from_browser: str = None) -> tuple:
    """Download audio from a URL using yt-dlp.

    Returns (filepath, title) on success or (None, error_message) on failure.
    """
    try:
        # Determine output filename
        if filename:
            safe_name = sanitize_filename(filename)
        else:
            # Try to get title from yt-dlp
            title_cmd = ["yt-dlp", "--no-download", "--print", "title",
                         "--no-warnings", url]
            if cookies_from_browser:
                title_cmd.extend(["--cookies-from-browser", cookies_from_browser])
            title_result = subprocess.run(
                title_cmd, capture_output=True, text=True, timeout=60
            )
            title = title_result.stdout.strip() if title_result.returncode == 0 else ""
            safe_name = sanitize_filename(title) if title else f"download_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}"

        output_template = str(download_dir / f"{safe_name}.%(ext)s")

        # Build download command
        download_cmd = [
            "yt-dlp",
            "--extract-audio",
            "--audio-format", "mp3",
            "--audio-quality", "0",
            "--output", output_template,
            "--no-warnings",
            "--no-playlist",
        ]
        if cookies_from_browser:
            download_cmd.extend(["--cookies-from-browser", cookies_from_browser])
        download_cmd.append(url)

        result = subprocess.run(
            download_cmd, capture_output=True, text=True, timeout=600
        )

        if result.returncode != 0:
            error = result.stderr.strip().split('\n')[-1] if result.stderr else "Unknown error"
            return None, f"yt-dlp failed: {error}"

        # Find the downloaded file
        expected = download_dir / f"{safe_name}.mp3"
        if expected.exists():
            return expected, safe_name

        # Fallback: look for any file with matching stem
        for f in download_dir.iterdir():
            if f.is_file() and f.stem == safe_name and f.suffix.lower() in AUDIO_VIDEO_EXTENSIONS:
                return f, safe_name

        return None, "Download completed but output file not found"

    except FileNotFoundError:
        return None, "yt-dlp not installed. Run: pip install yt-dlp"
    except subprocess.TimeoutExpired:
        return None, "Download timed out (10 minute limit)"
    except Exception as e:
        return None, str(e)


# ---------------------------------------------------------------------------
# Transcription
# ---------------------------------------------------------------------------

def transcribe_file(model, filepath: Path, output_dir: Path, language: str,
                    archive_dir: Path = None, keep_originals: bool = False,
                    metadata_header: str = None) -> tuple:
    """Transcribe a single file. Returns (success: bool, output_name or error)."""
    try:
        result = model.transcribe(str(filepath), fp16=False, language=language)
        txt_path = safe_output_path(output_dir, filepath.stem)
        txt_path.parent.mkdir(parents=True, exist_ok=True)

        with open(str(txt_path), "w", encoding="utf-8") as f:
            if metadata_header:
                f.write(metadata_header)
                f.write("\n---\n\n")
            f.write(result["text"])

        # Archive original
        if archive_dir and not keep_originals:
            dest = safe_archive_path(archive_dir, filepath.name)
            shutil.move(str(filepath), str(dest))

        return True, txt_path.name
    except Exception as e:
        return False, str(e).split('\n')[0]


def build_metadata_header(metadata: dict) -> str:
    """Build a readable header from Excel row metadata."""
    lines = []
    # Priority fields first
    priority = ["Touchplan", "Account", "Subject", "Owner ", "Status",
                "Primary Call Date/Time (UTC)", "Primary Call Duration",
                "Primary Call Type"]
    for key in priority:
        if key in metadata:
            label = key.strip()
            lines.append(f"{label}: {metadata[key]}")
    # Add any remaining fields not already included
    for key, val in metadata.items():
        if key not in priority and key not in ("Recording Link",
                "All Recording Links (chronological)", "SFDC Call Link", "f"):
            lines.append(f"{key.strip()}: {val}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()

    # Create output folder
    args.output.mkdir(parents=True, exist_ok=True)

    # Create input and archive folders if processing local files
    if args.input.is_dir():
        archive_dir = args.input / "Archive"
        archive_dir.mkdir(exist_ok=True)
    else:
        archive_dir = args.output / "Archive"
        archive_dir.mkdir(exist_ok=True)

    # ---------------------------------------------------------------
    # Phase 1: Collect all URLs (--url, links.txt, --xlsx)
    # ---------------------------------------------------------------
    url_jobs = []  # list of {url, name, metadata}

    # From --url flags
    for u in args.url:
        url_jobs.append({"url": u, "name": None, "metadata": {}})

    # From links.txt
    if args.input.is_dir():
        links_file_urls = read_links_file(args.input)
        if links_file_urls:
            print(f"Found {len(links_file_urls)} URL(s) in links.txt")
            for u in links_file_urls:
                url_jobs.append({"url": u, "name": None, "metadata": {}})

    # From Excel file
    if args.xlsx:
        if not args.xlsx.exists():
            print(f"Error: Excel file not found: {args.xlsx}")
            sys.exit(1)
        xlsx_entries = read_xlsx_urls(args.xlsx, args.url_column, args.name_columns)
        print(f"Found {len(xlsx_entries)} URL(s) in {args.xlsx.name}")
        url_jobs.extend(xlsx_entries)

    # Deduplicate URLs
    seen_urls = set()
    unique_jobs = []
    for job in url_jobs:
        if job["url"] not in seen_urls:
            seen_urls.add(job["url"])
            unique_jobs.append(job)
    url_jobs = unique_jobs

    # ---------------------------------------------------------------
    # Phase 2: Download all URLs
    # ---------------------------------------------------------------
    downloaded_files = []  # list of (Path, metadata_dict)
    if url_jobs:
        # Auto-detect browser for cookies if URL looks like Zoom CCI
        cookies_browser = args.cookies_from_browser
        has_zoom_cci = any("zoom.us/cci" in j["url"] for j in url_jobs)
        if has_zoom_cci and not cookies_browser:
            print("\nZoom Contact Center links detected. Trying browser cookies for authentication.")
            print("  (You must be logged into Zoom in your browser.)")
            print("  (To specify a browser, use --cookies-from-browser chrome|firefox|edge)\n")
            cookies_browser = "chrome"

        print(f"Downloading audio from {len(url_jobs)} URL(s)...\n")
        for i, job in enumerate(url_jobs, 1):
            url_display = job["url"][:80] + "..." if len(job["url"]) > 80 else job["url"]
            label = job["name"] or url_display
            print(f"  [{i}/{len(url_jobs)}] {label}")

            filepath, result = download_url_audio(
                job["url"],
                archive_dir.parent if archive_dir else args.output,
                filename=job["name"],
                cookies_from_browser=cookies_browser,
            )
            if filepath:
                size_mb = filepath.stat().st_size / (1024 * 1024)
                print(f"    Downloaded: {filepath.name} ({size_mb:.1f} MB)")
                downloaded_files.append((filepath, job.get("metadata", {})))
            else:
                print(f"    FAILED: {result}")

        # Archive links.txt if we read from it
        if args.input.is_dir():
            links_path = args.input / LINKS_FILENAME
            if links_path.exists():
                archive_links_file(args.input, archive_dir)

    # ---------------------------------------------------------------
    # Phase 3: Collect local media files
    # ---------------------------------------------------------------
    existing_transcripts = get_existing_transcripts(args.output)

    local_to_transcribe = []
    local_to_archive = []

    if args.input.is_dir():
        media_files = find_media_files(args.input)
        for f in media_files:
            # Skip files that were just downloaded (they're in downloaded_files)
            if any(dl[0].name == f.name for dl in downloaded_files):
                continue
            if transcript_exists(f.stem, existing_transcripts):
                local_to_archive.append(f)
            else:
                local_to_transcribe.append(f)

        # Archive already-transcribed local files
        if local_to_archive:
            print(f"\nAlready transcribed, archiving {len(local_to_archive)} file(s):")
            for f in local_to_archive:
                if not args.keep_originals:
                    dest = safe_archive_path(archive_dir, f.name)
                    shutil.move(str(f), str(dest))
                    print(f"  Archived: {f.name}")

    # Combine: downloaded files + local files to transcribe
    all_to_transcribe = []
    for filepath, metadata in downloaded_files:
        all_to_transcribe.append((filepath, metadata))
    for filepath in local_to_transcribe:
        all_to_transcribe.append((filepath, {}))

    if not all_to_transcribe:
        if url_jobs:
            print("\nNo files to transcribe. All downloads failed or already transcribed.")
        else:
            print(f"\nNo new files to transcribe in {args.input}")
        sys.exit(0)

    # ---------------------------------------------------------------
    # Phase 4: Transcribe
    # ---------------------------------------------------------------
    print(f"\n{len(all_to_transcribe)} file(s) to transcribe:")
    for filepath, _ in all_to_transcribe:
        size_mb = filepath.stat().st_size / (1024 * 1024)
        print(f"  - {filepath.name} ({size_mb:.1f} MB)")

    # Load Whisper
    try:
        import whisper
    except ImportError:
        print("\nError: openai-whisper not installed.")
        print("Run: pip install openai-whisper")
        sys.exit(1)

    print(f"\nLoading Whisper '{args.model}' model...")
    model = whisper.load_model(args.model)
    print("Model loaded.\n")

    total_start = time.time()
    results = {"success": [], "failed": []}

    for i, (filepath, metadata) in enumerate(all_to_transcribe, 1):
        if not filepath.exists():
            print(f"[{i}/{len(all_to_transcribe)}] Skipping (file missing): {filepath.name}")
            continue

        print(f"[{i}/{len(all_to_transcribe)}] Transcribing: {filepath.name} ...", flush=True)
        start = time.time()

        header = build_metadata_header(metadata) if metadata else None
        success, result_name = transcribe_file(
            model, filepath, args.output, args.language,
            archive_dir=archive_dir, keep_originals=args.keep_originals,
            metadata_header=header,
        )

        elapsed = time.time() - start
        if success:
            print(f"  Saved: {result_name} ({elapsed:.1f}s)")
            results["success"].append((filepath.name, result_name))
        else:
            print(f"  FAILED ({elapsed:.1f}s): {result_name}")
            results["failed"].append((filepath.name, result_name))

    # ---------------------------------------------------------------
    # Summary
    # ---------------------------------------------------------------
    total = time.time() - total_start
    print(f"\n{'='*50}")
    print(f"Done in {total:.1f}s")
    if downloaded_files:
        print(f"  URLs downloaded: {len(downloaded_files)}")
    if local_to_archive:
        print(f"  Already transcribed (archived): {len(local_to_archive)}")
    print(f"  Newly transcribed: {len(results['success'])}")
    print(f"  Failed: {len(results['failed'])}")

    if results["failed"]:
        print("\nFailed files:")
        for name, err in results["failed"]:
            print(f"  - {name}: {err}")

    print(f"\nTranscripts saved to: {args.output}")
    if not args.keep_originals:
        print(f"Originals archived to: {archive_dir}")


if __name__ == "__main__":
    main()
